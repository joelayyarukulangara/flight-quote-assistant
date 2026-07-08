"""
Export service: WhatsApp customer text, PDF customer quote, Excel comparison
sheet. Customer-facing exports never show API source, internal score, or
technical detail -- only clean route/date/airline/fare information.
"""
from datetime import datetime

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from utils.date_utils import format_datetime, now_str


class ExportError(Exception):
    pass


def build_whatsapp_text(quote_result, company_name="Your Travel Company"):
    request = quote_result.request
    currency = request.currency
    lines = []
    lines.append(f"{request.package_name} - Flight Options")
    lines.append("")
    lines.append(f"Customer: {request.customer_name}")
    lines.append(f"Package: {request.package_name}")
    passengers = request.passengers
    p_parts = [f"{passengers.adults} Adults"]
    if passengers.children:
        p_parts.append(f"{passengers.children} Children")
    if passengers.infants:
        p_parts.append(f"{passengers.infants} Infants")
    lines.append(f"Passengers: {' + '.join(p_parts)}")
    lines.append("")

    for i, combo in enumerate(quote_result.combinations, start=1):
        up, down = combo.up_flight, combo.down_flight
        lines.append(f"Option {i}: {combo.category}")
        lines.append(f"UP: {up.origin} to {up.destination}, {up.departure_datetime.strftime('%d %b')}, {up.airline}")
        lines.append(f"DOWN: {down.origin} to {down.destination}, {down.departure_datetime.strftime('%d %b')}, {down.airline}")
        lines.append(f"Fare: {currency} {combo.final_quote_per_person:,.0f} per person")
        lines.append("")

    lines.append(f"Fare checked on: {now_str()}")
    lines.append(quote_result.fare_warning)
    return "\n".join(lines)


def export_pdf(quote_result, file_path, company_name="Your Travel Company"):
    try:
        request = quote_result.request
        doc = SimpleDocTemplate(file_path, pagesize=A4, topMargin=20 * mm, bottomMargin=20 * mm)
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle("TitleStyle", parent=styles["Title"], fontSize=16)
        heading_style = ParagraphStyle("HeadingStyle", parent=styles["Heading2"], spaceBefore=10)
        normal = styles["Normal"]
        small = ParagraphStyle("Small", parent=styles["Normal"], fontSize=8, textColor=colors.grey)

        elements = [
            Paragraph(company_name, title_style),
            Paragraph(f"{request.package_name} - Flight Options", heading_style),
            Spacer(1, 6),
            Paragraph(f"Customer: {request.customer_name}", normal),
            Paragraph(f"Package: {request.package_name}", normal),
            Paragraph(
                f"Passengers: {request.passengers.adults} Adults, "
                f"{request.passengers.children} Children, {request.passengers.infants} Infants",
                normal,
            ),
            Spacer(1, 12),
        ]

        for i, combo in enumerate(quote_result.combinations, start=1):
            up, down = combo.up_flight, combo.down_flight
            elements.append(Paragraph(f"Option {i}: {combo.category}", heading_style))
            data = [
                ["", "From", "To", "Date", "Airline"],
                ["UP", up.origin, up.destination, up.departure_datetime.strftime("%d %b %Y"), up.airline],
                ["DOWN", down.origin, down.destination, down.departure_datetime.strftime("%d %b %Y"), down.airline],
            ]
            table = Table(data, colWidths=[35, 60, 60, 90, 110])
            table.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2b3a55")),
                        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                        ("FONTSIZE", (0, 0), (-1, -1), 9),
                    ]
                )
            )
            elements.append(table)
            elements.append(
                Paragraph(
                    f"Fare: {request.currency} {combo.final_quote_per_person:,.0f} per person "
                    f"(Total group fare: {request.currency} {combo.total_group_fare:,.0f})",
                    normal,
                )
            )
            elements.append(Spacer(1, 10))

        elements.append(Spacer(1, 10))
        elements.append(Paragraph(f"Fare checked on: {now_str()}", small))
        elements.append(Paragraph(quote_result.fare_warning, small))

        doc.build(elements)
    except Exception as exc:
        raise ExportError(f"Failed to export PDF: {exc}") from exc


def export_excel(quote_result, file_path):
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Quote Comparison"

        headers = [
            "Option Title", "Package Days/Nights", "UP Route", "UP Airline", "UP Departure",
            "UP Arrival", "UP Stops", "UP Fare", "DOWN Route", "DOWN Airline", "DOWN Departure",
            "DOWN Arrival", "DOWN Stops", "DOWN Fare", "Base Fare", "Markup",
            "Final Quote Per Person", "Total Group Fare", "Warnings",
        ]
        ws.append(headers)

        for combo in quote_result.combinations:
            up, down = combo.up_flight, combo.down_flight
            ws.append(
                [
                    combo.category,
                    f"{combo.package_days}D/{combo.package_nights}N",
                    f"{up.origin}-{up.destination}",
                    up.airline,
                    format_datetime(up.departure_datetime),
                    format_datetime(up.arrival_datetime),
                    up.stops,
                    up.fare,
                    f"{down.origin}-{down.destination}",
                    down.airline,
                    format_datetime(down.departure_datetime),
                    format_datetime(down.arrival_datetime),
                    down.stops,
                    down.fare,
                    combo.total_base_fare_per_person,
                    combo.markup_amount,
                    combo.final_quote_per_person,
                    combo.total_group_fare,
                    "; ".join(combo.warnings),
                ]
            )

        for i, _ in enumerate(headers, start=1):
            ws.column_dimensions[get_column_letter(i)].width = 20

        wb.save(file_path)
    except Exception as exc:
        raise ExportError(f"Failed to export Excel: {exc}") from exc
